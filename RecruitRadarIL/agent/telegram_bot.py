"""
RecruitRadar-IL Telegram control bot.

Runs on the researcher's own machine and is driven entirely from the Telegram
app on your phone. It uses long-polling - the script reaches OUT to Telegram's
servers - so it needs no public IP and no port forwarding. As long as this
machine is on and online, you can command it from anywhere.

The bot is only a remote control + inbox: the actual scoring runs here,
locally, against data/recruitradar.db. Scoring is rule- and statistics-driven
only (no LLM decides anything, no model is trained on message content).

Commands (also registered so Telegram shows hints):
  /scan          re-score the corpus and receive the new-leads CSV
  /top [N]       quick text preview of top N flagged messages (does NOT mark
                 them as sent - use for browsing without exhausting the queue)
  /proposals     pending channel-discovery proposals
  /approve NAME  approve a proposed channel (adds it to channels_extra.txt)
  /reject NAME   reject a proposed channel
  /status        summary of the last run
  /help          this message

Setup (once):
  1. In Telegram, message @BotFather -> /newbot -> follow prompts -> copy the token.
  2. Put it in agent/.env:   TELEGRAM_BOT_TOKEN=123456:ABC...
  3. Run:  python agent/telegram_bot.py
  4. Message your bot /start - it replies with your chat id. Put that in
     agent/.env as BOT_OWNER_ID=... and restart, so only you can drive it.

Everything it surfaces is a lead for review, not a conclusion.
"""

import io
import os
import sys
import time
import threading
from pathlib import Path
from datetime import datetime, timezone

import requests

ROOT = Path(__file__).resolve().parent.parent
os.chdir(ROOT)
sys.path.insert(0, str(ROOT / "agent"))
import pipeline  # noqa: E402  (engine; see agent/pipeline.py)


def _load_env(path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        if v.strip():
            os.environ.setdefault(k.strip(), v.strip())

_load_env(ROOT / ".env")
_load_env(ROOT / "agent" / ".env")

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
# Tolerate a value that was pasted with surrounding quotes or whitespace
# (a common mistake when saving the secret) - strip them before parsing.
OWNER = os.getenv("BOT_OWNER_ID", "").strip().strip('"').strip("'").strip()
OWNER = int(OWNER) if OWNER.lstrip("-").isdigit() else None
API = f"https://api.telegram.org/bot{TOKEN}"

_scan_lock = threading.Lock()

HELP = (
    "RecruitRadar-IL control bot\n\n"
    "/scan - re-score the corpus and get the new-leads CSV\n"
    "/top [N] - quick text preview of top N flagged (does not mark as sent)\n"
    "/proposals - pending channel proposals\n"
    "/approve NAME - approve a proposed channel\n"
    "/reject NAME - reject a proposed channel\n"
    "/status - last run summary\n"
    "/help - this message\n\n"
    "Leads for review, not conclusions."
)


def api(method, **params):
    try:
        r = requests.post(f"{API}/{method}", json=params, timeout=60)
        return r.json()
    except requests.RequestException as e:
        print(f"[api] {method} failed: {e}")
        return {"ok": False}


def send(chat_id, text):
    # Telegram hard-limits a message to 4096 chars; chunk on line boundaries.
    text = text or "(empty)"
    while text:
        chunk, text = text[:3900], text[3900:]
        if text and "\n" in chunk:
            cut = chunk.rfind("\n")
            text, chunk = chunk[cut:] + text, chunk[:cut]
        r = api("sendMessage", chat_id=chat_id, text=chunk)
        # Surface Telegram-side failures instead of dropping messages silently -
        # e.g. 429 rate limit, blocked chat, or a bad chat_id.
        if not r.get("ok"):
            print(f"[send] telegram rejected: {r.get('error_code')} "
                  f"{r.get('description')} (chat_id={chat_id}, "
                  f"chunk_len={len(chunk)})")


def send_document(chat_id, filename, content_bytes, caption=None,
                  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
    """Upload a file to the chat via sendDocument. Multipart, not JSON."""
    try:
        files = {"document": (filename, content_bytes, mime)}
        data = {"chat_id": chat_id}
        if caption:
            data["caption"] = caption
        r = requests.post(f"{API}/sendDocument", data=data, files=files, timeout=120)
        j = r.json()
    except requests.RequestException as e:
        print(f"[send_document] {e}")
        return {"ok": False}
    if not j.get("ok"):
        print(f"[send_document] telegram rejected: {j.get('error_code')} "
              f"{j.get('description')} (chat_id={chat_id}, filename={filename})")
    return j


def _snippet(s, n=90):
    return " ".join((s or "").split())[:n]


def fmt_leads_preview(leads):
    if not leads:
        return "No leads at or above p>=0.5 - run /scan first."
    out = [f"Top {len(leads)} flagged (preview, not marked as sent):\n"]
    for i, r in enumerate(leads, 1):
        out.append(f"{i}. p={r['p']:.3f}  [{r['channel']}]\n   {_snippet(r['text'])}")
    return "\n".join(out)


def fmt_proposals(props):
    if not props:
        return "No pending channel proposals."
    out = [f"{len(props)} pending proposals (approve with /approve NAME):\n"]
    for r in props:
        out.append(f"@{r['candidate']}  base={r['base_rate']:.2f} cent={r['centrality']:.2f}")
    return "\n".join(out)


def fmt_status(m):
    if not m:
        return "No run yet - send /scan."
    return (f"Last run {m['run_id']}\n"
            f"messages: {m['n_messages']} | channels: {m['n_channels']}\n"
            f"LF coverage: {m['lf_coverage']:.0%}\n"
            f"flagged total (p>=0.5): {m.get('n_flagged', '-')}\n"
            f"pending proposals: {m['n_proposals_pending']}\n"
            f"verdicts so far: {m['n_verdicts']}\n"
            f"label model: {m['label_model']}")


def build_digest_caption(summary, n_new):
    """Short caption that ships with the XLSX attachment."""
    return (f"RecruitRadar-IL digest\n"
            f"run {summary['run_id']}\n"
            f"{n_new} new leads (p>=0.5) since last digest\n"
            f"{summary['n_messages']} messages scanned across "
            f"{summary['n_channels']} channels\n"
            f"verdicts so far: {summary['n_verdicts']}")


# Column order optimised for scanning: row number first so a reader can
# refer to a specific lead by "#7"; then score (immediate risk cue); then
# when/where; then the Hebrew translation (populated when the original was
# Russian); then the original text; then trailing metadata.
REPORT_COLS = ["#", "p_recruitment", "date", "channel", "text_he",
               "text", "category", "rule_hits", "sender_hash", "msg_id"]

# Column widths (in Excel character units) tuned for phone-viewer readability.
# Text columns get generous width + wrap_text so long messages render as
# multi-line paragraphs inside the cell instead of an unreadable strip.
COL_WIDTHS = {
    "#": 5,               "p_recruitment": 8,   "date": 20,   "channel": 20,
    "text_he": 55,        "text": 55,   "category": 14,
    "rule_hits": 30,      "sender_hash": 18,   "msg_id": 10,
}


def build_xlsx_bytes(df):
    """Render the unsent-leads DataFrame as a styled XLSX workbook and return
    its bytes. The formatting is what makes this readable on a phone:

    * Frozen header row, bold header background, auto-filter enabled.
    * Wide text columns with wrap_text=True so long messages become paragraphs
      inside the cell instead of a single overflowing line.
    * Row colouring by p_recruitment - deep red for very high (>= 0.8), amber
      for the mid band, so the eye lands on the most suspicious rows first.
    * Column order: score, when, where, Hebrew, original, then trailing meta.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    df = df.copy().reset_index(drop=True)
    df["text_he"] = pipeline.translate_series(df["text"])
    # 1-based row number so the reader can say "look at lead #7" and everyone
    # opens the same row - Excel's own row numbers include the header row.
    df["#"] = range(1, len(df) + 1)
    for col in REPORT_COLS:
        if col not in df.columns:
            df[col] = ""
    df = df[REPORT_COLS]

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    high_fill   = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    mid_fill    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin        = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top    = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # Header row
    ws.append(REPORT_COLS)
    for i, col in enumerate(REPORT_COLS, 1):
        c = ws.cell(row=1, column=i)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 22

    # Data rows
    for _, row in df.iterrows():
        ws.append([str(row[c]) if row[c] is not None else "" for c in REPORT_COLS])

    # Row styling: fill by score, wrap all text cells, thin borders throughout.
    # p_recruitment lives in column 2 now (column 1 is the row number).
    p_col = REPORT_COLS.index("p_recruitment") + 1
    for r in range(2, ws.max_row + 1):
        try:
            p = float(ws.cell(row=r, column=p_col).value)
        except (TypeError, ValueError):
            p = 0.0
        row_fill = high_fill if p >= 0.8 else (mid_fill if p >= 0.5 else None)
        # Set row height so wrapped text has vertical breathing room without
        # blowing up too much. openpyxl won't compute this from wrap alone.
        ws.row_dimensions[r].height = 90
        for col_idx in range(1, len(REPORT_COLS) + 1):
            c = ws.cell(row=r, column=col_idx)
            c.alignment = wrap_top
            c.border = cell_border
            if row_fill is not None:
                c.fill = row_fill

    # Freeze header and enable filtering on all columns.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Format p_recruitment as 3-decimal number for a clean glance.
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=p_col)
        try:
            cell.value = float(cell.value)
            cell.number_format = "0.000"
        except (TypeError, ValueError):
            pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def deliver_digest(chat_id, summary):
    """Ship (caption + styled XLSX of unsent flagged leads) to `chat_id`.
    Marks every lead sent afterward. If there are no new leads, sends NOTHING
    and prints a note to the log."""
    fresh = pipeline.unsent_flagged()
    n_new = len(fresh)
    if n_new == 0:
        print(f"[digest] no new leads to send (chat={chat_id})")
        return 0
    caption = build_digest_caption(summary, n_new)
    filename = f"recruitradar-leads-{datetime.now(timezone.utc):%Y%m%d-%H%M}.xlsx"
    xlsx_bytes = build_xlsx_bytes(fresh)
    r = send_document(chat_id, filename, xlsx_bytes, caption=caption)
    if r.get("ok"):
        pairs = list(zip(fresh["channel"], fresh["msg_id"]))
        pipeline.mark_sent(pairs)
        print(f"[digest] delivered {n_new} leads (chat={chat_id}, file={filename}, "
              f"bytes={len(xlsx_bytes)})")
    return n_new


def do_scan(chat_id):
    if not _scan_lock.acquire(blocking=False):
        send(chat_id, "A scan is already running - hold on.")
        return
    try:
        send(chat_id, "Scanning...")
        summary = pipeline.run_pipeline()
        n_new = deliver_digest(chat_id, summary)
        if n_new == 0:
            # /scan initiated by a human deserves an explicit acknowledgement,
            # unlike the cloud cron where silence-when-nothing-new is intended.
            send(chat_id, "Scan complete - no NEW leads to send this run.\n\n" +
                 fmt_status(summary))
    except Exception as e:
        send(chat_id, f"Scan failed: {e.__class__.__name__}: {e}")
    finally:
        _scan_lock.release()


def handle(update):
    msg = update.get("message") or update.get("edited_message")
    if not msg or "text" not in msg:
        return
    chat_id = msg["chat"]["id"]
    text = msg["text"].strip()
    cmd, _, arg = text.partition(" ")
    cmd, arg = cmd.lower(), arg.strip()

    if cmd in ("/start", "start"):
        if OWNER is None:
            send(chat_id, f"Your chat id is {chat_id}.\nAdd BOT_OWNER_ID={chat_id} to "
                          f"agent/.env and restart the bot so only you can use it.\n\n" + HELP)
        elif chat_id == OWNER:
            send(chat_id, "Ready.\n\n" + HELP)
        else:
            send(chat_id, "This bot is private.")
        return

    if OWNER is None:
        send(chat_id, "Set BOT_OWNER_ID in agent/.env first (send /start to get your id).")
        return
    if chat_id != OWNER:
        return

    if cmd in ("/help", "/menu"):
        send(chat_id, HELP)
    elif cmd == "/scan":
        threading.Thread(target=do_scan, args=(chat_id,), daemon=True).start()
    elif cmd == "/top":
        n = int(arg) if arg.isdigit() else 10
        send(chat_id, fmt_leads_preview(pipeline.top_leads(min(n, 25),
                                                          min_p=pipeline.FLAG_THRESHOLD)))
    elif cmd in ("/proposals", "/discover"):
        send(chat_id, fmt_proposals(pipeline.list_proposals()))
    elif cmd == "/approve":
        if not arg:
            send(chat_id, "Usage: /approve channel_name")
        else:
            r = pipeline.approve_channel(arg)
            send(chat_id, f"Approved @{r} - it enters CHANNELS on the next collect."
                 if r else f"No pending proposal named '{arg}'.")
    elif cmd == "/reject":
        if not arg:
            send(chat_id, "Usage: /reject channel_name")
        else:
            r = pipeline.reject_channel(arg)
            send(chat_id, f"Rejected @{r}." if r else f"No pending proposal named '{arg}'.")
    elif cmd == "/status":
        send(chat_id, fmt_status(pipeline.last_run()))
    else:
        send(chat_id, "Unknown command. /help for the list.")


def main():
    if not TOKEN:
        sys.exit("TELEGRAM_BOT_TOKEN is not set. Create a bot with @BotFather, "
                 "then put the token in agent/.env (see agent/.env.example).")
    me = api("getMe")
    if not me.get("ok"):
        sys.exit("Telegram rejected the token. Double-check TELEGRAM_BOT_TOKEN in agent/.env.")
    print(f"Bot @{me['result']['username']} is up. "
          f"Owner: {OWNER if OWNER else 'UNSET - send /start to your bot to learn your id'}.")
    api("setMyCommands", commands=[
        {"command": "scan", "description": "re-score and get the new-leads CSV"},
        {"command": "top", "description": "quick preview of top flagged (no send-mark)"},
        {"command": "proposals", "description": "pending channel proposals"},
        {"command": "approve", "description": "approve a proposed channel"},
        {"command": "reject", "description": "reject a proposed channel"},
        {"command": "status", "description": "last run summary"},
        {"command": "help", "description": "command list"},
    ])

    offset = None
    print("Polling for commands (Ctrl+C to stop)...")
    while True:
        try:
            r = requests.get(f"{API}/getUpdates",
                             params={"offset": offset, "timeout": 50}, timeout=60)
            data = r.json()
        except requests.RequestException as e:
            print(f"[poll] {e}; retrying in 3s")
            time.sleep(3)
            continue
        if not data.get("ok"):
            time.sleep(3)
            continue
        for update in data["result"]:
            offset = update["update_id"] + 1
            try:
                handle(update)
            except Exception as e:
                print(f"[handle] {e.__class__.__name__}: {e}")


if __name__ == "__main__":
    main()
